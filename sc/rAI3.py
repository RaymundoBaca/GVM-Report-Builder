# rAI3.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def parse_summary(filename: Path) -> pd.DataFrame:
    with open(filename, "r", encoding="utf-8") as f:
        content = f.read()

    vulns = {}
    current_host = None
    current_vuln = None
    in_mitigation = False
    mitigation_lines = []

    lines = content.split('\n')

    for i, line in enumerate(lines):
        stripped_line = line.strip()

        # Detect host
        host_match = re.match(r"Host (\d+\.\d+\.\d+\.\d+)", stripped_line)
        if host_match:
            current_host = host_match.group(1)
            continue

        # Detect vulnerability + CVSS
        vuln_match = re.match(r"The following vulnerability is called NVT: (.+) \(CVSS: ([0-9.]+)\)", stripped_line)
        if vuln_match:
            vuln_name = vuln_match.group(1).strip()
            cvss_score = vuln_match.group(2).strip()

            current_vuln = vuln_name
            if current_vuln not in vulns:
                vulns[current_vuln] = {
                    "Mitigation": "",
                    "Hosts": set(),
                    "CVSS": float(cvss_score)
                }

            if current_host:
                vulns[current_vuln]["Hosts"].add(current_host)
            
            # Reset mitigation tracking
            in_mitigation = False
            mitigation_lines = []
            continue

        # Detect start of Mitigation section
        if stripped_line.startswith("Mitigation:"):
            in_mitigation = True
            mitigation_lines = []
            # Check if there's content on the same line after "Mitigation:"
            content_after = stripped_line[11:].strip()
            if content_after:
                mitigation_lines.append(content_after)
            continue

        # Collect mitigation text until empty line or next section
        if in_mitigation:
            if stripped_line == "" or stripped_line.startswith("Description:") or stripped_line.startswith("Impact:"):
                # End of mitigation section
                if current_vuln and mitigation_lines:
                    vulns[current_vuln]["Mitigation"] = " ".join(mitigation_lines).strip()
                in_mitigation = False
                mitigation_lines = []
            else:
                mitigation_lines.append(stripped_line)
            continue

        # Detect "already described"
        if "This vulnerability has already been described previously" in stripped_line:
            if current_vuln and current_host:
                vulns[current_vuln]["Hosts"].add(current_host)
            continue

        # Detect hosts listed directly (e.g., "['1.2.3.4', '5.6.7.8']")
        hosts_list_match = re.findall(r"\['([\d\., ]+)'\]", stripped_line)
        if hosts_list_match and current_vuln:
            host_ips = hosts_list_match[0].replace(" ", "").split(",")
            vulns[current_vuln]["Hosts"].update(host_ips)

    # Convert to DataFrame
    data = []
    for vuln, info in vulns.items():
        sorted_hosts = sorted(info["Hosts"], key=lambda ip: list(map(int, ip.split('.'))))
        data.append([
            vuln,
            info["Mitigation"],
            ", ".join(sorted_hosts),
            info["CVSS"]
        ])

    df = pd.DataFrame(
        data,
        columns=["Vulnerability found in Pentesting", "Mitigation", "Hosts", "CVSS"]
    )

    df = df.sort_values(by="CVSS", ascending=False).reset_index(drop=True)
    return df


def style_excel(file: Path) -> None:
    wb = load_workbook(file)
    ws = wb.active

    # General styles
    header_fill = PatternFill("solid", fgColor="3F6CAF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_style = Border(
        left=Side(border_style="thin", color="4C94D8"),
        right=Side(border_style="thin", color="4C94D8"),
        top=Side(border_style="thin", color="4C94D8"),
        bottom=Side(border_style="thin", color="4C94D8")
    )

    # Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = cell_center
        cell.border = border_style

    # Rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        vuln, mitig, hosts, cvss = row
        vuln.font = Font(name="Arial", size=11)
        mitig.font = Font(name="Arial", size=11)
        hosts.font = Font(name="Arial", size=10)
        try:
            score = float(cvss.value)
        except Exception:
            score = None

        if score is not None:
            cvss.font = Font(name="Arial", size=22, bold=True, color="FFFFFF")
            if 9.0 <= score <= 10.0:
                cvss.fill = PatternFill("solid", fgColor="EE0000")
            elif 7.0 <= score <= 8.9:
                cvss.fill = PatternFill("solid", fgColor="E97132")
            elif 4.0 <= score <= 6.9:
                cvss.fill = PatternFill("solid", fgColor="FFC000")

        for cell in row:
            cell.alignment = cell_center
            cell.border = border_style

    wb.save(file)


def main(pdf_path, txt_path, log, output_dir=None, progress_callback=None):
    """
    Generates the Excel table from Summary.txt.
    """
    if output_dir:
        results_dir = Path(output_dir) / "Results"
    else:
        results_dir = Path(pdf_path).parent / "Results"

    results_dir.mkdir(parents=True, exist_ok=True)

    input_file = results_dir / "Summary.txt"
    output_file = results_dir / "Table.xlsx"

    if not input_file.exists():
        log(f"⛔ File not found: {input_file}")
        return

    if progress_callback:
        progress_callback(0.0)

    try:
        df = parse_summary(input_file)
        if progress_callback:
            progress_callback(0.5)
    except Exception as e:
        log(f"⛔ Error parsing {input_file}: {e}")
        return

    try:
        if progress_callback:
            progress_callback(0.7)
        df.to_excel(output_file, index=False)
        if progress_callback:
            progress_callback(0.85)
        style_excel(output_file)
        if progress_callback:
            progress_callback(1.0)
        log(f"✅ Table generated: {output_file}")
    except Exception as e:
        log(f"⛔ Error generating Excel: {e}")
